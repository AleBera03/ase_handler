import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
import sys

# configuration
if len(sys.argv) <= 1:
    print("File is missing")
    exit
input_file = sys.argv[1] # 0: name of py program; 1: first actual arg
if (re.match(".*\\.s$", input_file)):
    base_name, suffix = input_file.split(".")
else:
    print("File format is wrong (only .s)")
    exit
output_file = f"{base_name}.xlsx" 
ncc = 100           

# read every instruction
istructions = []
label = ""
with open(input_file, "r") as f:
    for line in f:
        line = line.strip()
        line = label + line
        label = ""
        if "#" in line:
            line = line.split("#", 1)[0].strip()
        if (re.search("^([^:])*:$", line)):
            label = line + " "
            continue
        if line != "":
            istructions.append(line)
            

# create empty pandas dataframe 
cols = ["Conteggio"] + ["Disassemblato"] + [f"{i+1}" for i in range(ncc)]
df = pd.DataFrame("", index=range(len(istructions)), columns=cols)
df["Disassemblato"] = istructions

# export pandas to excel
df.to_excel(output_file, index=False)

# set colors
colors = {
    "F": "00FF0000",
    "D": "00FF9900",  
    "E": "00FFFF00",  
    "M": "0000FF00", 
    "W": "0000FFFF",  
    "S": "00808080"
}

# reopen file
wb = load_workbook(output_file)
ws = wb.active

# set styles
lightgrayFill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
grayFill = PatternFill(start_color="00808080", end_color="00808080", fill_type="solid")
outBorder = Border(left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
    )

# modify first two column
ws.column_dimensions["B"].width = 20
start_row = 2
end_row = len(istructions) + 1
for cell in ws["A"]:
    cell.fill = lightgrayFill
    cell.border = outBorder
    cell.alignment = Alignment(horizontal='center')
ws.column_dimensions["A"].width = 12
max_length = 0
for cell in ws["B"]:
    cell.fill = grayFill
    cell.font = Font(bold = True, name="Courier New")
    cell.border = outBorder
    try:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    except:
        pass
adjusted_width = (max_length + 2) * 1.2     # pretty accurate for monospace font case
ws.column_dimensions["B"].width = adjusted_width
ws.freeze_panes = ws["C2"]

# calculate the writing area in string format
start_col_letter = get_column_letter(3)
end_col_letter = get_column_letter(2 + ncc)
rangestr = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"

# restyle the writing area
i = 0
for row in ws[rangestr]:
    for cell in row:
        if (i%2 == 0):
            cell.fill = lightgrayFill
        cell.border = outBorder
        cell.alignment = Alignment(horizontal='center')
    i += 1

# conditional formatting
# FETCH
red_fill = PatternFill(start_color=colors["F"], end_color=colors["F"], fill_type="solid")
dxff = DifferentialStyle(fill=red_fill)
rule = Rule(type="containsText", operator="containsText", text="F", dxf=dxff)
rule.formula = ['NOT(ISERROR(SEARCH("F",C2)))']
ws.conditional_formatting.add(rangestr, rule)
# DECODE
orange_fill = PatternFill(start_color=colors["D"], end_color=colors["D"], fill_type="solid")
dxfd = DifferentialStyle(fill=orange_fill)
rule = Rule(type="containsText", operator="containsText", text="D", dxf=dxfd)
rule.formula = ['NOT(ISERROR(SEARCH("D",C2)))']
ws.conditional_formatting.add(rangestr, rule)
# EXE
yellow_fill = PatternFill(start_color=colors["E"], end_color=colors["E"], fill_type="solid")
dxfe = DifferentialStyle(fill=yellow_fill)
rule = Rule(type="containsText", operator="containsText", text="E", dxf=dxfe)
rule.formula = ['NOT(ISERROR(SEARCH("E",C2)))']
ws.conditional_formatting.add(rangestr, rule)
# MEM
green_fill = PatternFill(start_color=colors["M"], end_color=colors["M"], fill_type="solid")
dxfm = DifferentialStyle(fill=green_fill)
rule = Rule(type="containsText", operator="containsText", text="M", dxf=dxfm)
rule.formula = ['NOT(ISERROR(SEARCH("M",C2)))']
ws.conditional_formatting.add(rangestr, rule)
# WRITE BACK
lblue_fill = PatternFill(start_color=colors["W"], end_color=colors["W"], fill_type="solid")
dxfw = DifferentialStyle(fill=lblue_fill)
rule = Rule(type="containsText", operator="containsText", text="W", dxf=dxfw)
rule.formula = ['NOT(ISERROR(SEARCH("W",C2)))']
ws.conditional_formatting.add(rangestr, rule)


# write formulas within first column
for i in range(2, end_row):
    formula = f"=IF(COUNTIF({start_col_letter}{i}:{end_col_letter}{i},\"W\")=0, -1, INDIRECT(ADDRESS(1, MAX(IF({start_col_letter}{i}:{end_col_letter}{i}=\"W\", COLUMN({start_col_letter}{i}:{end_col_letter}{i}), 0)), 4)))"
    cell_string = f"A{i}"
    ws[cell_string].value = ArrayFormula(ref=cell_string, text=formula)


wb.save(output_file)
wb.close()